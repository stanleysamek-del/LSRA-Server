from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import os
from io import BytesIO
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont

app = Flask(__name__)
CORS(app)

# Paths
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "LSRA_TEMPLATE.xlsx")
LOGO_PATH = os.path.join(os.path.dirname(__file__), "ASHE_logo.jpg")

@app.route("/")
def index():
    return jsonify({"status": "ok", "message": "LSRA server running"})

@app.route("/generate", methods=["POST"])
def generate_lsra():
    try:
        data = request.get_json(force=True)
        print("🔹 Incoming LSRA request:", data)

        if not os.path.exists(TEMPLATE_PATH):
            return jsonify({"error": "Template not found"}), 500

        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb["Tool"]

        # Insert ASHE logo at A1
        if os.path.exists(LOGO_PATH):
            try:
                img = Image(LOGO_PATH)
                img.width, img.height = 120, 40
                ws.add_image(img, "A1")
                print("🖼️ ASHE logo inserted at A1")
            except Exception as e:
                print("⚠️ Logo insertion failed:", e)

        # Rich text content for A15
        rt = CellRichText()
        rt.append(TextBlock(InlineFont(b=True, rFont="Calibri", sz=1100), "Date: "))
        rt.append(TextBlock(InlineFont(i=True, rFont="Calibri", sz=1100), f"{data.get('dateOfInspection','')}\n"))

        rt.append(TextBlock(InlineFont(b=True, rFont="Calibri", sz=1100), "Location Address: "))
        rt.append(TextBlock(InlineFont(i=True, rFont="Calibri", sz=1100), f"{data.get('address','')}\n"))

        rt.append(TextBlock(InlineFont(b=True, rFont="Calibri", sz=1100), "Action(s) Taken: "))
        rt.append(TextBlock(InlineFont(i=True, rFont="Calibri", sz=1100),
            "Creation of Corrective Action Plan, ILSM created, notified engineering.\n"))

        rt.append(TextBlock(InlineFont(b=True, rFont="Calibri", sz=1100), "Person Completing Life Safety Risk Matrix: "))
        rt.append(TextBlock(InlineFont(i=True, rFont="Calibri", sz=1100), f"{data.get('inspector','')}\n"))

        rt.append(TextBlock(InlineFont(b=True, rFont="Calibri", sz=1100), "ILSM Required? "))
        rt.append(TextBlock(InlineFont(rFont="Calibri", sz=1100), "YES"))

        ws["A15"].rich_text = rt
        ws["A15"].alignment = Alignment(wrap_text=True, vertical="top")

        # Save workbook in memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"LSRA_{data.get('facilityName','Facility').replace(' ','_')}_{data.get('floorName','Floor').replace(' ','_')}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print("❌ LSRA generation failed:", e)
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
